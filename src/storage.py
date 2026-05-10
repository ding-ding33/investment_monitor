import os
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import pandas as pd

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data")
XLSX_PATH = os.path.join(DATA_DIR, "portfolio_history.xlsx")
LEGACY_TSV_PATH = os.path.join(DATA_DIR, "portfolio_history.tsv")

COLUMNS = [
    "date", "asset_class", "fund_code", "fund_name",
    "shares", "nav", "nav_date", "market_value", "cost_basis",
    "pnl_pct", "weight_pct", "nav_source", "total_value", "total_pnl_pct",
]


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """统一列结构，确保历史读取和追加写入兼容。"""
    return df.reindex(columns=COLUMNS)


def _read_history_df() -> pd.DataFrame:
    """优先读取 xlsx；若仅存在旧 tsv，自动兼容导入。"""
    if os.path.exists(XLSX_PATH):
        df = pd.read_excel(XLSX_PATH, sheet_name="history")
        return _normalize_columns(df)
    if os.path.exists(LEGACY_TSV_PATH):
        df = pd.read_csv(LEGACY_TSV_PATH, sep="\t")
        return _normalize_columns(df)
    return pd.DataFrame(columns=COLUMNS)


def _write_history_xlsx(df: pd.DataFrame) -> None:
    os.makedirs(DATA_DIR, exist_ok=True)
    with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="history")
        ws = writer.sheets["history"]
        ws.freeze_panes = "A2"
        for idx, col in enumerate(COLUMNS, start=1):
            col_letter = get_column_letter(idx)
            values = [col] + ["" if pd.isna(v) else str(v) for v in df[col].tolist()]
            max_len = max(len(v) for v in values)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 36)
            for cell in ws[col_letter]:
                cell.alignment = Alignment(horizontal="left", vertical="center")


def _row_has_values(row: pd.Series) -> bool:
    for value in row.tolist():
        if not pd.isna(value) and str(value).strip() != "":
            return True
    return False


def save_snapshot(snapshot: dict):
    """将当日快照按长格式追加写入 Excel。"""
    today = snapshot["date"]
    total_value = snapshot["total_value"]
    total_pnl = snapshot["total_pnl_pct"]

    rows = []
    for snap in snapshot["positions"] + [snapshot["cash"]]:
        rows.append({
            "date": today,
            "asset_class": snap["asset_class"],
            "fund_code": snap["fund_code"],
            "fund_name": snap["fund_name"],
            "shares": snap["shares"],
            "nav": snap["nav"] if snap["nav"] is not None else None,
            "nav_date": snap["nav_date"] or None,
            "market_value": snap["market_value"],
            "cost_basis": snap["cost_basis"],
            "pnl_pct": snap["pnl_pct"],
            "weight_pct": snap["weight_pct"],
            "nav_source": snap.get("nav_source", "live"),
            "total_value": total_value,
            "total_pnl_pct": total_pnl,
        })

    history_df = _read_history_df()
    new_df = pd.DataFrame(rows, columns=COLUMNS)
    if not history_df.empty and _row_has_values(history_df.iloc[-1]):
        spacer = pd.DataFrame([{c: None for c in COLUMNS}], columns=COLUMNS)
        history_df = pd.concat([history_df, spacer], ignore_index=True)

    merged = pd.concat([history_df, new_df], ignore_index=True)
    _write_history_xlsx(merged)
    print(f"\n数据已写入: {XLSX_PATH}")


def read_last_navs() -> dict[str, dict]:
    """从历史表格中读取上一日各项资产净值，作为兜底数据。

    Returns:
        {fund_code: {"nav": float, "nav_date": str}}
    """
    df = _read_history_df()
    if df.empty:
        return {}

    valid = df[df["date"].notna()]
    valid = valid[valid["date"].astype(str).str.strip() != ""]
    if valid.empty:
        return {}

    last_date = str(valid["date"].iloc[-1])
    last_rows = valid[valid["date"].astype(str) == last_date]

    result = {}
    for _, r in last_rows.iterrows():
        code = str(r["fund_code"])
        nav = r["nav"]
        if code == "CASH" or pd.isna(nav) or str(nav).strip() == "":
            continue
        nav_date = "" if pd.isna(r["nav_date"]) else str(r["nav_date"])
        result[code] = {"nav": float(nav), "nav_date": nav_date}

    return result
